import os
import urllib.parse
from typing import Optional, List
from datetime import date, datetime

from fastapi import APIRouter, Depends, Query, Security, UploadFile
from fastapi.responses import FileResponse
from openpyxl.styles.numbers import NumberFormat
from sqlalchemy.orm import Session

from app.utils.jwt import token_user
from app.utils.common_utils import order_numbering, log_msg
from app.utils.date_utils import D
from app.utils.pg_kcp import cancel as kcp_cancel, CancelData, ResCancelData
from app.utils.pg_payco import cancel as payco_cancel, ReqCancel, ResCancel
from app.utils.sms_util import send_sms_customer
from app.utils.ecoupon_m12 import info, ReqInfo, ResInfo, cancel as m12_cancel, ResCancel as m12ResCancel, ReqExtend, extend
from app.utils.ecoupon_dnc import cancel as dnc_cancel, ReqCancel as DncReqCancel, ResCancel as DncResCancel, info as dnc_info, ReqInfo as DncReqInfo, ResInfo as DncResInfo
from app.utils.ecoupon_kt import EcounponKt, ReqCancel as KtReqCancel, ResBase as KtResBase
from app.errors import exceptions as exc

from app.database.conn import db
from app.database.schema import Order, OrderProduct, PgInfo, OrderShipping, OrderRe, PgCancel, OrderReProduct, ProductOption, LogOrder, Customer, Store, PgInfoSub, AdminRefund, OptionValue

from app.models.common import Success, LogOrderDataIn, ListLog, SuccessCount
from app.models.order import ListDataOrderProduct, DataOrderProduct, DataOrder, ModOrder, ModShipping, ListDataOrderRe, DataOrderRe, CancelPartOrder, ModRe, DataPgCancel, ModOrderStatus, ListAdminRefund
from app.models.auth import MemberToken

from app.common.consts import PG_KCP_CANCEL_ALL, PG_KCP_CANCEL_PART, ORDER_STATUS
from app.common.config import conf

router = APIRouter(prefix='/order')


@router.get("", response_model=ListDataOrderProduct, name="주문 목록")
def list_order(store_code: Optional[str] = Query(default=None, description="상점 코드"),
               store_name: Optional[str] = Query(default=None, description="상점 명"),
               status: Optional[str] = Query(default=None, description="상태"),
               pg_type: Optional[str] = Query(default=None, description="결제 수단"),
               pg_sub_type: Optional[str] = Query(default=None, description="추가 결제 수단"),
               prd_code: Optional[str] = Query(default=None, description="상품 코드"),
               prd_type: Optional[str] = Query(default=None, description="상품 타입"),
               prd_name: Optional[str] = Query(default=None, description="상품명"),
               pa_id: Optional[int] = Query(default=None, description="PA id"),
               order_id: Optional[int] = Query(default=None, description="order id"),
               order_name: Optional[str] = Query(default=None, description="주문자명"),
               order_email: Optional[str] = Query(default=None, description="주문자 id(Email)"),
               s_reg_date: Optional[date] = Query(default=None, description="등록일 시작"),
               e_reg_date: Optional[date] = Query(default=None, description="등록일 종료"),
               offset: int = 0, limit: int = 20,
               session: Session = Depends(db.session),
               user: MemberToken = Security(token_user, scopes=["read:order"])):
    qry = (session.query(OrderProduct, Order, PgInfo)
           .join(Order, Order.id == OrderProduct.order_id)
           .join(Store, Order.store_code == Store.code)
           .outerjoin(PgInfo, PgInfo.order_id == OrderProduct.order_id)
           .outerjoin(PgInfoSub, PgInfoSub.pg_info_id == PgInfo.id).filter(Order.status != "D"))

    if store_code:
        qry = qry.filter(Order.store_code == store_code)

    if store_name:
        qry = qry.filter(Store.title.like(f"%{store_name}%"))

    if status:
        qry = qry.filter(OrderProduct.status == status)

    if pg_type:
        if pg_type == 'payco':
            qry = qry.filter(PgInfo.provider == 'PAYCO')
            if pg_sub_type:
                qry = qry.filter(PgInfoSub.kind == pg_sub_type)
        else:
            qry = qry.filter(PgInfo.kind == pg_type)

    if prd_code:
        qry = qry.filter(OrderProduct.product_code == prd_code)

    if prd_type:
        qry = qry.filter(OrderProduct.type.like(f"{prd_type}%"))

    if prd_name:
        qry = qry.filter(OrderProduct.product_name.like(f"%{prd_name}%"))

    if pa_id:
        qry = qry.filter(OrderProduct.member_id == pa_id)

    if order_name:
        qry = qry.filter(Order.user_name.like(f"%{order_name}%"))

    if order_email:
        qry = qry.filter(Order.user_email == order_email)

    if order_id:
        qry = qry.filter(OrderProduct.order_id == order_id)
    else:
        if s_reg_date and e_reg_date:
            qry = qry.filter(Order.reg_date.between(s_reg_date, D().make235959(e_reg_date)))
        elif s_reg_date:
            qry = qry.filter(Order.reg_date > s_reg_date)
        elif e_reg_date:
            qry = qry.filter(Order.reg_date < D().make235959(e_reg_date))

    qry = qry.group_by(OrderProduct.id)
    qry = qry.order_by(Order.reg_date.desc())

    total = qry.count()
    qry = qry.offset(offset).limit(limit)
    datas = qry.all()

    order_products = []

    for data in datas:
        dop = DataOrderProduct.from_orm(data[0])
        dop.order_date = data[1].reg_date
        if data[2]:
            dop.pg_date = datetime.strptime(data[2].app_time, "%Y%m%d%H%M%S")
            dop.pg_kind = data[2].kind
        if data[1].origin_order_id:
            dop.origin_order_id = data[1].origin_order_id
        dop.user_name = data[1].user_name
        dop.store_name = data[1].store.title
        dop.store_code = data[1].store_code
        dop.product_option_name = get_option_name(data[0])
        order_products.append(dop)

    return ListDataOrderProduct(total=total, datas=order_products)


@router.get("/excel", name="주문 내역 엑셀 다운로드")
def excel_order(store_code: Optional[str] = Query(default=None, description="상점 코드"),
                status: Optional[str] = Query(default=None, description="상태"),
                pg_type: Optional[str] = Query(default=None, description="결제 수단"),
                pg_sub_type: Optional[str] = Query(default=None, description="추가 결제 수단"),
                prd_code: Optional[str] = Query(default=None, description="상품 코드"),
                prd_type: Optional[str] = Query(default=None, description="상품 타입"),
                prd_name: Optional[str] = Query(default=None, description="상품명"),
                pa_id: Optional[int] = Query(default=None, description="PA id"),
                order_id: Optional[int] = Query(default=None, description="order id"),
                order_name: Optional[str] = Query(default=None, description="주문자명"),
                order_email: Optional[str] = Query(default=None, description="주문자 id(Email)"),
                s_reg_date: Optional[date] = Query(default=None, description="등록일 시작"),
                e_reg_date: Optional[date] = Query(default=None, description="등록일 종료"),
                session: Session = Depends(db.session),
                user: MemberToken = Security(token_user, scopes=["read:order"])):
    qry = (session.query(OrderProduct, Order, PgInfo)
           .join(Order, Order.id == OrderProduct.order_id)
           .outerjoin(PgInfo, PgInfo.order_id == OrderProduct.order_id)
           .outerjoin(PgInfoSub, PgInfoSub.pg_info_id == PgInfo.id).filter(Order.status != "D"))

    if store_code:
        qry = qry.filter(Order.store_code == store_code)

    if status:
        qry = qry.filter(OrderProduct.status == status)

    if pg_type:
        if pg_type == 'payco':
            qry = qry.filter(PgInfo.provider == 'PAYCO')
            if pg_sub_type:
                qry = qry.filter(PgInfoSub.kind == pg_sub_type)
        else:
            qry = qry.filter(PgInfo.kind == pg_type)

    if prd_code:
        qry = qry.filter(OrderProduct.product_code == prd_code)

    if prd_type:
        qry = qry.filter(OrderProduct.type.like(f"{prd_type}%"))

    if prd_name:
        qry = qry.filter(OrderProduct.product_name.like(f"%{prd_name}%"))

    if pa_id:
        qry = qry.filter(OrderProduct.member_id == pa_id)

    if order_id:
        qry = qry.filter(OrderProduct.order_id == order_id)

    if order_name:
        qry = qry.filter(Order.user_name.like(f"%{order_name}%"))

    if order_email:
        qry = qry.filter(Order.user_email == order_email)

    if s_reg_date and e_reg_date:
        qry = qry.filter(Order.reg_date.between(s_reg_date, D().make235959(e_reg_date)))
    elif s_reg_date:
        qry = qry.filter(Order.reg_date > s_reg_date)
    elif e_reg_date:
        qry = qry.filter(Order.reg_date < D().make235959(e_reg_date))

    qry = qry.group_by(OrderProduct.id)
    qry = qry.order_by(Order.reg_date.desc())

    total = qry.count()
    datas = qry.all()

    order_products = []

    for data in datas:
        dop = DataOrderProduct.from_orm(data[0])
        dop.order_date = data[1].reg_date
        if data[2]:
            dop.pg_date = datetime.strptime(data[2].app_time, "%Y%m%d%H%M%S")
            dop.pg_kind = data[2].kind
        if data[1].origin_order_id:
            dop.origin_order_id = data[1].origin_order_id
        dop.user_name = data[1].user_name
        dop.user_mobile = data[1].user_mobile
        dop.store_name = data[1].store.title
        dop.store_code = data[1].store_code
        if dop.type == 'DP':
            dop.recipient_name = data[1].recipient_name
            dop.recipient_mobile = data[1].recipient_mobile
        else:
            dop.recipient_name = data[0].user_name
            dop.recipient_mobile = data[0].user_phone
        dop.zipcode = data[1].zipcode
        dop.address = data[1].address
        dop.address_detail = data[1].address_detail
        dop.shipping_msg = data[1].shipping_msg
        dop.product_option_name = get_option_name(data[0])
        order_products.append(dop)

    import openpyxl
    file_name = f'ConiaOrderList-{D.now_str_trim()}.xlsx'
    file_path = os.path.join(os.getcwd(), 'app/upload/', file_name)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["상품유형", "주문일시", "주문번호", "주문자명", "주문자연락처", "공급사", "상점명", "상품명", "옵션명", "수량", "금액", "할인금액", "결제금액", "결제일시", "구매확정일", "결제수단", "상태", "수령인", "수령인연락처", "우편번호", "배송지", "배송지 상세", "배송메모"])

    for row in order_products:
        final_amount = row.amount
        if row.discount:
            final_amount -= row.discount
        sheet.append(
            [row.type, row.pg_date, row.order_id, row.user_name, row.user_mobile, row.seller_title, row.store_name, row.product_name, row.product_option_name, row.ea, row.amount, row.discount, final_amount, row.pg_date, row.complete_date, row.pg_kind, ORDER_STATUS.get(row.status, ''), row.recipient_name, row.recipient_mobile, row.zipcode,
             row.address, row.address_detail, row.shipping_msg])
    wb.save(file_path)

    return FileResponse(file_path, filename=file_name)


@router.put("/update_status", response_model=Success, name="주문 상태 일괄 수정")
def update_status(indata: ModOrderStatus,
                  session: Session = Depends(db.session),
                  user: MemberToken = Security(token_user, scopes=["write:order"])):
    target_ids = list(set(indata.order_product_ids))

    for order_product_id in target_ids:
        op: OrderProduct = OrderProduct.get(session=session, id=order_product_id)
        if not op:
            continue
        status = indata.status
        if op.status != status:
            # 로깅
            log_data = LogOrderDataIn(action="주문상태 일괄수정", order_id=op.order_id, msg=log_msg("msg", f"{op.status} -> {status}"), writer=f"{user.name}:{user.id}")
            LogOrder.create(session, auto_commit=True, **log_data.dict())
            if status == 'DW':
                OrderProduct.filter(session, order_shipping_id=op.order_shipping_id).update_q(status='DW')
            else:
                op.status = status
            session.commit()

    return Success()


@router.get("/admin-refund", response_model=ListAdminRefund, name="관리자 환불 내역")
def list_admin_refund(order_id: Optional[str] = Query(default=None, description="주문 번호"),
                      product_name: Optional[str] = Query(default=None, description="상품명"),
                      store_name: Optional[str] = Query(default=None, description="상점명"),
                      store_code: Optional[str] = Query(default=None, description="상점 코드"),
                      customer_email: Optional[str] = Query(default=None, description="주문자 이메일"),
                      customer_name: Optional[str] = Query(default=None, description="주문자 이름"),
                      s_reg_date: Optional[date] = Query(default=None, description="등록일 시작"),
                      e_reg_date: Optional[date] = Query(default=None, description="등록일 종료"),
                      offset: int = 0, limit: int = 20,
                      session: Session = Depends(db.session),
                      user: MemberToken = Security(token_user, scopes=["read:order"])):
    qry = session.query(AdminRefund).join(Customer, Customer.id == AdminRefund.customer_id)

    if order_id:
        qry = qry.filter(AdminRefund.order_id == order_id)

    if product_name:
        qry = qry.filter(AdminRefund.product_name.like(f"%{product_name}%"))

    if store_name:
        qry = qry.filter(AdminRefund.store_name.like(f"%{store_name}%"))

    if store_code:
        qry = qry.filter(AdminRefund.store_code == store_code)

    if customer_email:
        qry = qry.filter(Customer.email.like(f"%{customer_email}%"))

    if customer_name:
        qry = qry.filter(Customer.name.like(f"%{customer_name}%"))

    if s_reg_date and e_reg_date:
        qry = qry.filter(AdminRefund.reg_date.between(s_reg_date, D().make235959(e_reg_date)))
    elif s_reg_date:
        qry = qry.filter(AdminRefund.reg_date > s_reg_date)
    elif e_reg_date:
        qry = qry.filter(AdminRefund.reg_date < D().make235959(e_reg_date))

    total = qry.count()
    qry = qry.order_by(AdminRefund.reg_date.desc())
    datas = qry.offset(offset).limit(limit).all()

    return ListAdminRefund(total=total, datas=datas)


@router.get("/same-shipping-cnt", response_model=int, name="주문 배송 동일 건수 확인")
def same_shipping_cnt(order_product_id: int,
                      session: Session = Depends(db.session),
                      user: MemberToken = Security(token_user, scopes=["read:order"])):
    op: OrderProduct = OrderProduct.get(session=session, id=order_product_id)
    return session.query(OrderProduct).filter(OrderProduct.order_shipping_id == op.order_shipping_id).count()


@router.post("/exchange", response_model=Success, name="교환 주문 생성")
def exchange(order_product_id: str,
             exchange_product_option_id: str,
             amount: int,
             count: int,
             session: Session = Depends(db.session),
             user: MemberToken = Security(token_user, scopes=["write:order"])):
    op: OrderProduct = OrderProduct.get(session=session, id=order_product_id)
    exchange_product: ProductOption = ProductOption.get(session=session, id=exchange_product_option_id)
    if not op:
        raise exc.NotFoundDataEx

    origin_order = op.order
    origin_shipping = op.order_shipping

    exchange_order = Order()
    exchange_order.id = order_numbering()
    exchange_order.origin_order_id = origin_order.id
    exchange_order.customer_id = origin_order.customer_id
    exchange_order.store_code = origin_order.store_code

    exchange_order.origin_amount = 0
    exchange_order.raw_amount = 0
    exchange_order.final_amount = 0
    # exchange_order.tex_free_amount =
    # exchange_order.tax_rate =
    exchange_order.discount = 0
    exchange_order.coupon_discount = 0
    exchange_order.status = "DW"

    exchange_order.user_name = origin_order.user_name
    exchange_order.user_phone = origin_order.user_phone
    exchange_order.user_mobile = origin_order.user_mobile
    exchange_order.user_email = origin_order.user_email

    exchange_order.recipient_name = origin_order.recipient_name
    exchange_order.recipient_phone = origin_order.recipient_phone
    exchange_order.recipient_mobile = origin_order.recipient_mobile
    exchange_order.zipcode = origin_order.zipcode
    exchange_order.address = origin_order.address
    exchange_order.address_detail = origin_order.address_detail
    exchange_order.shipping_msg = origin_order.shipping_msg

    exchange_order.shipping_cost = 0
    exchange_order.shipping_cost_post = 0
    exchange_order.shipping_condition = ""

    exchange_order.step_type = "exchange"
    exchange_order.client_type = ""
    exchange_order.referer = ""
    exchange_order.referer_url = ""
    exchange_order.total_ea = count
    exchange_order.total_kind = 1
    # exchange_order.ipcc_code = ""
    exchange_order.ip = ""

    session.add(exchange_order)
    session.commit()

    os = OrderShipping()
    os.status = 'R'
    os.cost = 0
    os.shipping_type = origin_shipping.shipping_type
    os.pay_type = origin_shipping.pay_type
    os.order_id = exchange_order.id
    os.member_id = origin_shipping.member_id
    os.shipping_info_id = origin_shipping.shipping_info_id
    session.add(os)
    session.commit()

    op = OrderProduct()
    op.status = "DW"
    op.order_id = exchange_order.id
    op.product_option_id = exchange_product.id
    op.product_code = exchange_product.product.code
    op.product_id = exchange_product.product_id
    op.ea = count
    op.product_name = exchange_product.product.name
    op.product_thumbnail = exchange_product.product.photos[0].uri
    op.amount = amount
    op.origin_price = exchange_product.origin_price * count
    op.seller_title = exchange_product.product.member.company.name if exchange_product.product.member.company.name else exchange_product.product.member.name

    op.member_id = exchange_product.product.member.id
    op.product_description = ''
    op.order_shipping_id = os.id
    op.type = exchange_product.product.type

    session.add(op)

    exchange_product.inventory.count = exchange_product.inventory.count - count
    session.commit()

    # 로깅
    log_data = LogOrderDataIn(action="교환주문 생성", order_id=str(origin_order.id), msg=log_msg("msg", f"교환주문 - {exchange_order.id}"), writer=f"{user.name}:{user.id}")
    LogOrder.create(session, auto_commit=True, **log_data.dict())

    return Success()


@router.get("/bulk_excel_shipping", name="송장번호 일괄등록용 엑셀파일 생성")
def make_excel_shipping(session: Session = Depends(db.session),
                        user: MemberToken = Security(token_user, scopes=["write:order"])):
    is_cm = False
    if "CM" in user.member_class or user.admin == 'Y':
        is_cm = True

    if is_cm:
        target = session.query(OrderProduct).filter(OrderProduct.status == 'DW').all()
    else:
        target = session.query(OrderProduct).filter(OrderProduct.status == 'DW', OrderProduct.member_id == user.id).all()

    order_shipping_ids = {}
    for op in target:
        if op.order_shipping_id:
            ods = order_shipping_ids.get(op.order_shipping_id)
            if ods:
                ods.append(op)
            else:
                order_shipping_ids[op.order_shipping_id] = [op]

    import openpyxl
    from openpyxl.styles import Alignment, Font

    file_name = f'ConiaShippingList-{D.now_str_trim()}.xlsx'
    file_path = os.path.join(os.getcwd(), 'app/upload/', file_name)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["주문번호", "배송UID", "상품정보", "택배사", "송장번호", "", "", "택배사 목록"])

    for k, v in order_shipping_ids.items():
        prd_infos = []

        for prd in v:
            opt_name = get_option_name(prd)
            if not opt_name:
                opt_name = '-'
            prd_infos.append(f"상품명 : {prd.product_name} | 옵션명 : {opt_name} | 수량 : {prd.ea}개")
        prd_info = '\n'.join(prd_infos)

        sheet.append([str(v[0].order_id), str(k), prd_info])

    logistics = OptionValue.filter(session, type='logistics').all()
    logistics_idx = 0
    for i, v in enumerate(logistics, start=2):
        sheet.cell(row=i, column=8, value=v.name)
        logistics_idx = i
    font = Font(color="FF0000", bold=True)
    info_cell = sheet.cell(row=logistics_idx+1, column=8)
    info_cell.value = '※ 송장 번호는 " -, / " 등 기호를 제외하고 숫자만 입력해 주세요.'
    info_cell.font = font

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.number_format = '@'
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['H'].width = 60

    wb.save(file_path)

    return FileResponse(file_path, filename=file_name)


@router.put("/bulk_excel_shipping", response_model=SuccessCount, name="송장번호 일괄등록(Excel)")
async def bulk_excel_shipping(file: UploadFile,
                              session: Session = Depends(db.session),
                              user: MemberToken = Security(token_user, scopes=['write:order'])):
    allowed_extensions = {'xlsx'}
    if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
        raise exc.NotAllowedFileEx

    file_path = os.path.join(os.getcwd(), 'app/upload/', f'ConiaShippingBulk-{D.now_str_trim()}.xlsx')

    with open(file_path, 'wb') as buffer:
        buffer.write(await file.read())

    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    is_cm = False
    if "CM" in user.member_class or user.admin == 'Y':
        is_cm = True

    success_cnt = 0
    try:
        logistic_dict = {}
        logistics = OptionValue.filter(session, type='logistics').all()
        for logistic in logistics:
            logistic_dict[logistic.name] = logistic.value
        for row in ws.rows:
            row_a = row[0].value
            if row_a == '주문번호':
                continue
            if not row[1].value or not row[3].value or not row[4].value or not logistic_dict[row[3].value]:
                continue
            if not row_a:
                break
            if is_cm:
                order_shipping = OrderShipping.get(session, id=row[1].value, status='R')
            else:
                order_shipping = OrderShipping.get(session, id=row[1].value, status='R', member_id=user.id)
            if not order_shipping:
                continue

            ship_out(session, order_shipping, row[4].value, row[3].value, logistic_dict[row[3].value], user, is_bulk=True)
            success_cnt += 1

        session.commit()
    finally:
        try:
            wb.close()
            os.remove(file_path)
        except:
            pass

    return SuccessCount(count=success_cnt)


@router.put("/shipping/{order_shipping_id}", response_model=Success, name="출고 처리")
def ship_process(order_shipping_id: int,
                 indata: ModShipping,
                 session: Session = Depends(db.session),
                 user: MemberToken = Security(token_user, scopes=["write:order"])):
    data: OrderShipping = OrderShipping.get(session=session, id=order_shipping_id)
    if not data:
        raise exc.NotFoundDataEx

    ship_out(session, data, indata.delivery_code, indata.delivery_provider, indata.delivery_provider_code, user)

    return Success()


@router.put("/shipping/{order_shipping_id}/force_end", response_model=Success, name="배송완료 강제변경")
def order_shipping_force_end(order_shipping_id: int,
                             session: Session = Depends(db.session),
                             user: MemberToken = Security(token_user, scopes=["write:order"])):
    os: OrderShipping = OrderShipping.get(session=session, id=order_shipping_id)
    if not os:
        raise exc.NotFoundDataEx

    os.update(session=session, auto_commit=True, status="DC", complete_date=D().now)
    OrderProduct.filter(session=session, order_shipping_id=order_shipping_id, status='DN').update_q(True, status="DC")

    # 로깅
    log_data = LogOrderDataIn(action="배송완료 강제변경", order_id=os.order_id, msg=log_msg("msg", f"배송완료 강제변경"), writer=f"{user.name}:{user.id}")
    LogOrder.create(session, auto_commit=True, **log_data.dict())

    return Success()


@router.get("/re", response_model=ListDataOrderRe, name="주문 반품 교환 요청 목록")
def list_re(store_code: Optional[str] = Query(default=None, description="상점 코드"),
            status: Optional[str] = Query(default=None, description="상태"),
            pa_id: Optional[int] = Query(default=None, description="PA id"),
            order_id: Optional[int] = Query(default=None, description="order id"),
            order_name: Optional[str] = Query(default=None, description="주문자 명"),
            order_email: Optional[str] = Query(default=None, description="주문자 id(Email)"),
            s_reg_date: Optional[date] = Query(default=None, description="등록일 시작"),
            e_reg_date: Optional[date] = Query(default=None, description="등록일 종료"),
            offset: int = 0, limit: int = 20,
            session: Session = Depends(db.session),
            user: MemberToken = Security(token_user, scopes=["read:order"])):
    qry = session.query(OrderReProduct).join(OrderProduct, OrderProduct.id == OrderReProduct.order_product_id).join(Order, Order.id == OrderProduct.order_id)

    if store_code:
        qry = qry.filter(Order.store_code == store_code)

    if status:
        qry = qry.filter(OrderProduct.status == status)

    if pa_id:
        qry = qry.filter(OrderProduct.member_id == pa_id)

    if order_id:
        qry = qry.filter(OrderProduct.order_id == order_id)

    if order_name:
        qry = qry.filter(Order.user_name.like(f"%{order_name}%"))

    if order_email:
        qry = qry.filter(Order.user_email == order_email)

    if s_reg_date and e_reg_date:
        qry = qry.filter(OrderReProduct.reg_date.between(s_reg_date, D().make235959(e_reg_date)))
    elif s_reg_date:
        qry = qry.filter(OrderReProduct.reg_date > s_reg_date)
    elif e_reg_date:
        qry = qry.filter(OrderReProduct.reg_date < D().make235959(e_reg_date))

    qry = qry.order_by(OrderReProduct.reg_date.desc())

    total = qry.count()
    qry = qry.offset(offset).limit(limit)
    datas = qry.all()

    return ListDataOrderRe(total=total, datas=datas)


@router.get("/re/{re_id}", response_model=DataOrderRe, name="주문 반품 교환 요청 상세")
def get_re(re_id: int,
           session: Session = Depends(db.session),
           user: MemberToken = Security(token_user, scopes=["read:order"])):
    data: OrderRe = OrderRe.get(session=session, id=re_id)
    if not data:
        raise exc.NotFoundDataEx

    return data


@router.put("/re/{re_id}", response_model=Success, name="주문 반품 교환 상태 변경")
def mod_re(re_id: int,
           indata: ModRe,
           session: Session = Depends(db.session),
           user: MemberToken = Security(token_user, scopes=["write:order"])):
    data: OrderRe = OrderRe.get(session=session, id=re_id)
    if not data:
        raise exc.NotFoundDataEx

    for order_product_id in indata.order_product_ids:
        orp = OrderReProduct.get(session, order_re_id=re_id, order_product_id=order_product_id)
        orp.product.update(session, True, status=indata.status)
        if indata.status == 'EXC' or indata.status == 'RFC':
            orp.update(session, True, end_date=D().now)

        # 로깅
        log_data = LogOrderDataIn(action="반품교환 상태변경", order_id=data.order_id, msg=log_msg("msg", f"{orp.product.product_name} / {indata.status}"), writer=f"{user.name}:{user.id}")
        LogOrder.create(session, auto_commit=True, **log_data.dict())

    return Success()


@router.get("/{order_id}", response_model=DataOrder, name="주문 상세")
def get_order(order_id: int,
              session: Session = Depends(db.session),
              user: MemberToken = Security(token_user, scopes=["read:order"])):
    order_data: Order = Order.get(session=session, id=order_id)
    if not order_data:
        raise exc.NotFoundDataEx

    cancel_disable = None
    if order_data.pg_info and order_data.pg_info.app_time[:8] != D().yyyymmdd() and order_data.pg_info.pg_info_sub:
        sub_list: List[PgInfoSub] = order_data.pg_info.pg_info_sub
        for sub in sub_list:
            if sub.kind == '식권 쿠폰':
                cancel_disable = '식권 쿠폰 사용 결제 입니다.'
                break

    result: DataOrder = DataOrder.from_orm(order_data)
    if cancel_disable:
        result.pg_cancel_disable = cancel_disable
    ops: List[OrderProduct] = OrderProduct.filter(session, order_id=order_id).all()

    order_products = []

    for data in ops:
        dop = DataOrderProduct.from_orm(data)
        dop.order_date = order_data.reg_date
        if order_data.pg_info:
            dop.pg_date = datetime.strptime(order_data.pg_info.app_time, "%Y%m%d%H%M%S")
            dop.pg_kind = order_data.pg_info.kind
            if cancel_disable:
                dop.pg_cancel_disable = cancel_disable

        dop.store_name = order_data.store.title
        dop.store_code = order_data.store_code
        dop.product_option_name = get_option_name(data)

        # E쿠폰 상품 사용전 이거나 잔액이 남았을 경우 쿠폰 상태 확인
        if data.ecoupon and (data.status == 'PD' or data.balance > 0):
            if data.ecoupon.provider == 'M12':
                req = ReqInfo(tr_id=data.ecoupon.tr_id, goods_id=data.ecoupon.goods_id, pin_no=data.ecoupon.pin_code)
                res: ResInfo = info(req)

                if res.PIN_STATUS != 'R':
                    if res.PIN_STATUS == 'F':
                        data.status = 'CP'
                        dop.status = 'CP'
                        if res.GOODS_TYPE == '04':
                            data.balance = int(res.ACCOUNT_BALANCE)
                            dop.balance = int(res.ACCOUNT_BALANCE)
                        if res.USE_DATE:
                            if len(res.USE_DATE) == 8:
                                data.complete_date = datetime.strptime(res.USE_DATE, "%Y%m%d")
                            elif len(res.USE_DATE) == 12:
                                data.complete_date = datetime.strptime(res.USE_DATE, "%Y%m%d%H%M")
                            elif len(res.USE_DATE) == 14:
                                data.complete_date = datetime.strptime(res.USE_DATE, "%Y%m%d%H%M%S")
                    elif res.PIN_STATUS == 'C':
                        data.status = 'CD'
                        dop.status = 'CD'
                    elif res.PIN_STATUS == 'V':
                        data.status = 'EXP'
                        dop.status = 'EXP'
                    session.commit()
            elif data.ecoupon.provider == 'DNC':
                req = DncReqInfo(pinCode=data.ecoupon.pin_code, orderCode=data.ecoupon.order_code)
                try:
                    res: DncResInfo = dnc_info(req)

                    print(res)
                except Exception as e:
                    if isinstance(e, exc.APIException):
                        print(e.detail)
                    else:
                        print(e.args)
                    pass

        order_products.append(dop)

    result.products = order_products

    return result


@router.put("/{order_id}", response_model=Success, name="주문 수정")
def mod_order(order_id: int, indata: ModOrder,
              session: Session = Depends(db.session),
              user: MemberToken = Security(token_user, scopes=["write:order"])):
    data: Order = Order.get(session=session, id=order_id)
    if not data:
        raise exc.NotFoundDataEx

    change_data = data.update_optional(session=session, auto_commit=True, **indata.dict())

    # 로깅
    log_data = LogOrderDataIn(action="수정", order_id=order_id, msg=log_msg("list", change_data), writer=f"{user.name}:{user.id}")
    LogOrder.create(session, auto_commit=True, **log_data.dict())

    return Success()


@router.get("/{order_id}/cancel", response_model=List[DataPgCancel], name="주문 취소 내역")
def order_cancel(order_id: str,
                 session: Session = Depends(db.session),
                 user: MemberToken = Security(token_user, scopes=["read:order"])):
    cancel_info: List[DataPgCancel] = PgCancel.filter(session, pg_info_order_id=order_id).all()
    return cancel_info


@router.post("/{order_id}/cancel", response_model=Success, name="주문 전체 취소")
def cancel(order_id: str,
           session: Session = Depends(db.session),
           user: MemberToken = Security(token_user, scopes=["write:order"])):
    order: Order = Order.get(session, id=order_id)

    if not order:
        raise exc.NotFoundDataEx

    if order.status == "PW":
        order.update(session, True, status="CD")
        OrderProduct.filter(session, order_id=order_id).update_q(True, status='CD')
    else:
        pg_info = PgInfo.get(session, order_id=order_id)

        cancel_time = ""
        cancel_tno = ""

        if pg_info.provider == "KCP":
            c = conf()
            data = CancelData(site_cd=c.PG_KCP_SITE_CODE,
                              kcp_cert_info=c.PG_KCP_CERT_INFO,
                              mod_type=PG_KCP_CANCEL_ALL,
                              tno=pg_info.tid)

            res: ResCancelData = kcp_cancel(data)
            cancel_time = res.canc_time
            cancel_tno = res.tno
        elif pg_info.provider == "PAYCO":
            data = ReqCancel(orderNo=pg_info.tid,
                             cancelTotalAmt=str(int(order.final_amount)),
                             orderCertifyKey=pg_info.cancel_key)
            res: ResCancel = payco_cancel(data)
            cancel_time = res.cancelYmdt
            cancel_tno = res.cancelTradeSeq

        order.update(session, True, status="CD")
        OrderProduct.filter(session, order_id=order_id).update_q(True, status='CD')
        ops: List[OrderProduct] = session.query(OrderProduct).filter(OrderProduct.order_id == order_id).all()
        for op in ops:
            if op.ecoupon:
                if op.ecoupon.provider == 'M12':
                    m12_req = ReqInfo(tr_id=op.ecoupon.tr_id, goods_id=op.ecoupon.goods_id, pin_no=op.ecoupon.pin_code)
                    m12_res: m12ResCancel = m12_cancel(m12_req)
                    if m12_res.STATUS_CODE == '0000':
                        op.ecoupon.status = 'D'
                    else:
                        op.ecoupon.status = 'DE'
                elif op.ecoupon.provider == 'DNC':
                    dnc_req = DncReqCancel(tradeId=op.ecoupon.trade_id, pinCode=op.ecoupon.pin_code, orderCode=op.ecoupon.order_code, requestedAt=D().now_str())
                    try:
                        res: DncResCancel = dnc_cancel(dnc_req)
                        if res.pinCode:
                            op.ecoupon.status = 'D'
                        else:
                            op.ecoupon.status = 'DE'
                    except:
                        op.ecoupon.status = 'DE'
                elif op.ecoupon.provider == 'KT':
                    kt_req = KtReqCancel(tradeId=op.ecoupon.tr_id, pinCode=op.ecoupon.pin_code)
                    try:
                        ec_kt = EcounponKt()
                        res: KtResBase = ec_kt.cancel(kt_req)
                        if res.resCode == '0000':
                            op.ecoupon.status = 'D'
                        else:
                            op.ecoupon.status = 'DE'
                    except:
                        op.ecoupon.status = 'DE'

        pg_info.update(session, True, cancel_type='all', cancel_date=datetime.strptime(cancel_time, "%Y%m%d%H%M%S"), cancel_mny=pg_info.amount, remain_amount=0)
        cancel_data = PgCancel()
        cancel_data.pg_info_order_id = order_id
        cancel_data.tno = cancel_tno
        cancel_data.type = "ALL"
        cancel_data.reg_date = datetime.strptime(cancel_time, "%Y%m%d%H%M%S")
        cancel_data.amount = pg_info.amount
        session.add(cancel_data)
        session.commit()

        # 로깅
        log_data = LogOrderDataIn(action="주문 전체취소", order_id=order_id, msg=log_msg("msg", f"주문 전체취소"), writer=f"{user.name}:{user.id}")
        LogOrder.create(session, auto_commit=True, **log_data.dict())

    return Success()


@router.post("/{order_id}/cancel_part", response_model=Success, name="주문 부분 취소")
def cancel_part(order_id: str,
                indata: CancelPartOrder,
                session: Session = Depends(db.session),
                user: MemberToken = Security(token_user, scopes=["write:order"])):
    cancel_amount = indata.cancel_amount
    target_products = []
    target_shippings = []

    for order_product_id in indata.order_product_ids:
        op: OrderProduct = OrderProduct.get(session, id=order_product_id, order_id=order_id)

        if not op:
            raise exc.NotFoundDataEx
        discount = 0
        if op.discount:
            discount = op.discount
        cancel_amount = cancel_amount - (op.amount - discount)
        target_products.append(op)

    for order_shipping_id in indata.order_shipping_ids:
        os: OrderShipping = OrderShipping.get(session, id=order_shipping_id, order_id=order_id)

        if not os:
            raise exc.NotFoundDataEx

        cancel_amount = cancel_amount - os.cost
        target_shippings.append(os)

    if cancel_amount != 0:
        raise exc.BadRequestEx

    pg_info = PgInfo.get(session, order_id=order_id)
    if not pg_info:
        raise exc.NotFoundDataEx

    cancel_time = ""
    cancel_tno = ""

    if pg_info.provider == "KCP":
        c = conf()
        data = CancelData(site_cd=c.PG_KCP_SITE_CODE,
                          kcp_cert_info=c.PG_KCP_CERT_INFO,
                          mod_mny=int(indata.cancel_amount),
                          rem_mny=pg_info.remain_amount,
                          mod_type=PG_KCP_CANCEL_PART,
                          mod_desc="주문 취소",
                          tno=pg_info.tid)

        res: ResCancelData = kcp_cancel(data)
        cancel_time = res.canc_time
        cancel_tno = res.tno
    elif pg_info.provider == "PAYCO":
        data = ReqCancel(orderNo=pg_info.tid,
                         cancelTotalAmt=str(int(indata.cancel_amount)),
                         orderCertifyKey=pg_info.cancel_key)
        res: ResCancel = payco_cancel(data)
        cancel_time = res.cancelYmdt
        cancel_tno = res.cancelTradeSeq

    remain_amount = pg_info.remain_amount - int(indata.cancel_amount)
    pg_info.update(session, True,
                   cancel_type='part',
                   cancel_date=datetime.strptime(cancel_time, "%Y%m%d%H%M%S"),
                   remain_amount=remain_amount,
                   cancel_mny=int(indata.cancel_amount))

    pg_cancel = PgCancel()
    pg_cancel.pg_info_order_id = order_id
    pg_cancel.tno = cancel_tno
    pg_cancel.type = "PART"
    pg_cancel.amount = int(indata.cancel_amount)
    pg_cancel.remain = remain_amount
    if pg_info.provider == "KCP":
        pg_cancel.part_seq = res.mod_pacn_seq_no
    session.add(pg_cancel)
    session.commit()

    # 로깅
    log_data = LogOrderDataIn(action="주문 부분취소", order_id=order_id, msg=log_msg("msg", f"주문 부분취소 - 취소금액 : {indata.cancel_amount}"), writer=f"{user.name}:{user.id}")
    LogOrder.create(session, auto_commit=True, **log_data.dict())

    for row in target_products:
        row.update(session, status="CD")

        if row.ecoupon:
            if row.ecoupon.provider == 'M12':
                m12_req = ReqInfo(tr_id=row.ecoupon.tr_id, goods_id=row.ecoupon.goods_id, pin_no=row.ecoupon.pin_code)
                m12_res: m12ResCancel = m12_cancel(m12_req)
                if m12_res.STATUS_CODE == '0000':
                    row.ecoupon.status = 'D'
                else:
                    row.ecoupon.status = 'DE'
            elif row.ecoupon.provider == 'DNC':
                dnc_req = DncReqCancel(tradeId=row.ecoupon.trade_id, pinCode=row.ecoupon.pin_code, orderCode=row.ecoupon.order_code, requestedAt=D().now_str())
                try:
                    res: DncResCancel = dnc_cancel(dnc_req)
                    if res.pinCode:
                        row.ecoupon.status = 'D'
                    else:
                        row.ecoupon.status = 'DE'
                except:
                    row.ecoupon.status = 'DE'
            elif op.ecoupon.provider == 'KT':
                kt_req = KtReqCancel(tradeId=op.ecoupon.tr_id, pinCode=op.ecoupon.pin_code)
                try:
                    ec_kt = EcounponKt()
                    res: KtResBase = ec_kt.cancel(kt_req)
                    if res.resCode == '0000':
                        op.ecoupon.status = 'D'
                    else:
                        op.ecoupon.status = 'DE'
                except:
                    op.ecoupon.status = 'DE'

        # 로깅
        log_data = LogOrderDataIn(action="주문 부분취소", order_id=order_id, msg=log_msg("msg", f"주문 부분취소 - 상품 : {row.product_name} / {row.product_id}"), writer=f"{user.name}:{user.id}")
        LogOrder.create(session, auto_commit=True, **log_data.dict())

    for row in target_shippings:
        row.update(session, status="CD")
    session.commit()

    count = OrderProduct.filter(session, order_id=order_id, status__ne="CD").count()
    if count == 0:
        Order.filter(session, id=order_id).update_q(True, status="CD")

    return Success()


@router.put("/{order_id}/admin-refund", name="관리자 환불")
def admin_refund(order_id: str,
                 amount: int = Query(description='환불 금액'),
                 session: Session = Depends(db.session),
                 user: MemberToken = Security(token_user, scopes=["write:order"])):
    order: Order = Order.get(session, id=order_id)

    if not order:
        raise exc.NotFoundDataEx

    order.update(session, True, status="AR")
    op_list: List[OrderProduct] = OrderProduct.filter(session, order_id=order_id).all()
    OrderProduct.filter(session, order_id=order_id).update_q(True, status='AR')

    prd_cnt = len(op_list)
    cnt_str = ''
    if prd_cnt > 1:
        cnt_str = f" 외 {prd_cnt}건"

    ar_data = AdminRefund()
    ar_data.order_id = order_id
    ar_data.product_name = f"{op_list[0].product.name}{cnt_str}"
    ar_data.amount = amount
    ar_data.store_code = order.store_code
    ar_data.store_name = order.store.title
    ar_data.member_id = user.id
    ar_data.customer_id = order.customer_id
    session.add(ar_data)
    session.commit()

    # 로깅
    log_data = LogOrderDataIn(action="관리자 환불", order_id=order.id, msg=log_msg("msg", f"관리자 환불 {amount}원"), writer=f"{user.name}:{user.id}")
    LogOrder.create(session, auto_commit=True, **log_data.dict())

    return Success()


@router.get("/{order_id}/log", response_model=ListLog, name="주문 Log")
def order_log(order_id: str,
              offset: int = 0, limit: int = 20,
              session: Session = Depends(db.session),
              user: MemberToken = Security(token_user, scopes=['write:order'])):
    qry = session.query(LogOrder).filter(LogOrder.order_id == order_id, LogOrder._del != "Y")

    total = qry.count()
    qry = qry.order_by(LogOrder.reg_date.desc())
    datas = qry.offset(offset).limit(limit).all()

    return ListLog(total=total, datas=datas)


@router.put("/{order_id}/period/{order_product_id}", name="주문 상품 만료일 변경")
def order_period_update(order_id: str,
                        order_product_id: int,
                        use_end_date: date,
                        session: Session = Depends(db.session),
                        user: MemberToken = Security(token_user, scopes=["write:order"])):
    op: OrderProduct = OrderProduct.get(session, id=order_product_id, order_id=order_id)
    ori_date = op.use_end_date

    if not op:
        raise exc.NotFoundDataEx

    if op.ecoupon and op.ecoupon.provider == 'M12':
        req_data = ReqExtend(tr_id=op.ecoupon.tr_id, goods_id=op.ecoupon.goods_id, pin_no=op.ecoupon.pin_code, period=use_end_date.strftime('%Y%m%d'))
        try:
            res_data = extend(req_data)
            if res_data.STATUS_CODE == "0000":
                op.update(session, False, use_end_date=D().generate235959(use_end_date))
                op.ecoupon.period_date = D().generate235959(use_end_date)
            else:
                raise exc.PgFailEx(msg=res_data.RESULT_REASON, code=res_data.STATUS_CODE)
        except Exception as e:
            msg = ''
            if isinstance(e, exc.APIException) and e.detail:
                msg = urllib.parse.unquote(e.detail, encoding='euc-kr')

            raise exc.ProcessFailEx(msg=f"M12 쿠폰 연장 API 실패 {msg}")
    else:
        op.update(session, False, use_end_date=D().generate235959(use_end_date))
    if op.status == 'EXP':
        op.status = 'PD'
    session.commit()

    # 로깅
    log_data = LogOrderDataIn(action="만료일 변경", order_id=order_id, msg=log_msg("msg", f"만료일 변경 - {op.product_name}({op.id})\n{ori_date.strftime('%Y%m%d')} -> {use_end_date.strftime('%Y%m%d')}"), writer=f"{user.name}:{user.id}")
    LogOrder.create(session, auto_commit=True, **log_data.dict())

    return Success()


@router.put("/{order_id}/force_end/{order_product_id}", name="주문 상품 구매확정 강제변경")
def order_product_force_end(order_id: str,
                            order_product_id: int,
                            session: Session = Depends(db.session),
                            user: MemberToken = Security(token_user, scopes=["write:order"])):
    op: OrderProduct = OrderProduct.get(session, id=order_product_id, order_id=order_id)

    if not op:
        raise exc.NotFoundDataEx

    op.status = 'CP'
    op.complete_date = D().now
    session.commit()

    ops = session.query(OrderProduct).filter(OrderProduct.order_id == order_id, OrderProduct.status != "CP").all()
    if len(ops) == 0:
        op.order.status = 'CP'
        session.commit()

    # 로깅
    log_data = LogOrderDataIn(action="구매확정 강제변경", order_id=order_id, msg=log_msg("msg", f"구매확정 강제변경"), writer=f"{user.name}:{user.id}")
    LogOrder.create(session, auto_commit=True, **log_data.dict())

    return Success()


def get_option_name(order_prd: OrderProduct):
    option_name = []
    po = order_prd.product_option.__dict__
    for key, value in po.items():
        if value and len(key) == 8 and key[:7] == "option_" and key[-1].isdigit():
            option_name.append(value)
    return "/".join(option_name)


def ship_out(session: Session, order_shipping: OrderShipping, code, provider, provider_code, user: MemberToken, is_bulk=False):
    order_shipping.update(session=session, auto_commit=False, status="DN", number=code, provider=provider, provider_code=provider_code, number_reg_date=D().now)
    OrderProduct.filter(session=session, order_shipping_id=order_shipping.id, status='DW').update_q(False, status="DN")
    order_shipping.order.status = "DN"
    session.commit()

    customer = Customer.get(session=session, id=order_shipping.order.customer_id)
    store = Store.get(session=session, code=order_shipping.order.store_code)

    ops = OrderProduct.filter(session=session, order_shipping_id=order_shipping.id).all()
    prd_titles = []
    for op in ops:
        prd_titles.append(op.product_name)
    if len(prd_titles) > 1:
        prd_title = f"{prd_titles[0]} 외 {len(prd_titles) - 1}건"
    else:
        prd_title = prd_titles[0]

    msg = f"""안녕하세요 [{customer.name}]님!
    [{store.title}]입니다.

    고객님께서 기다리시던 소중한 상품이 출발하였습니다~!
    조금만 기다려주시면 주말, 공휴일 제외 2~3일 내로 상품이 배송될 예정입니다:)

    ■상품명: {prd_title}
    ■주문번호: {order_shipping.order_id}
    ■택배회사: {provider}
    ■운송장번호: {code}

    오늘도 기분 좋은 하루 보내세요!!
    """
    send_sms_customer(session, customer, store, msg)

    # 로깅
    action_title = "출고 처리(일괄)" if is_bulk else "출고 처리"
    log_data = LogOrderDataIn(action="출고 처리", order_id=order_shipping.order_id, msg=log_msg("msg", f"{provider} / {code}"), writer=f"{user.name}:{user.id}")
    LogOrder.create(session, auto_commit=True, **log_data.dict())
